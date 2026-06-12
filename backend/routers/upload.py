"""File upload + smart import router"""
import os, io, uuid, json, logging, hashlib, base64, re, time as _time, tempfile
from pathlib import Path

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from cryptography.fernet import Fernet

from config import UPLOAD_DIR, LLM_ENCRYPTION_KEY
from database import async_session
from models import (
    User, KnowledgeEntry, KnowledgeCategory, DailyQuestion,
    LLMProvider, EntryStatusEnum, SourceTypeEnum, ContentTypeEnum,
)
from schemas import ApiResponse
from auth import require_admin, get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()
ALLOWED_EXT = {".pdf", ".docx", ".xlsx", ".jpg", ".jpeg", ".png", ".mp4", ".mp3", ".wav", ".webm"}


def _safe_filename(raw_name: str) -> str:
    """
    修复 Windows 上 GBK 中文文件名被 Starlette 误当 Latin-1 解码的问题。
    "ÏúÊÛ¾­Àí¼Ý..." → "销售经理驾驶..."
    如果文件名正常则原样返回。
    """
    if not raw_name:
        return raw_name
    # 试探：所有字节是否在 cp1252 可逆范围内
    try:
        raw_bytes = raw_name.encode("cp1252")
    except UnicodeEncodeError:
        # 已含非 Latin-1 字符 → UTF-8 文件名，正常
        return raw_name
    # 用 GBK 解码这些字节 → 命中的话就是中文恢复成功
    try:
        decoded = raw_bytes.decode("gbk")
        cn_count = sum(1 for c in decoded if '一' <= c <= '鿿')
        if cn_count >= 2:
            logger.info(f"_safe_filename 恢复中文: {raw_name!r} -> {decoded!r}")
            return decoded
    except UnicodeDecodeError:
        pass
    return raw_name

_ocr_engine = None
def _get_ocr():
    global _ocr_engine
    if _ocr_engine is None:
        from rapidocr_onnxruntime import RapidOCR
        _ocr_engine = RapidOCR()
    return _ocr_engine


@router.post("/upload/file", response_model=ApiResponse)
async def upload_file(file: UploadFile = File(...), _admin: User = Depends(require_admin)):
    safe_name_raw = _safe_filename(file.filename)
    suffix = Path(safe_name_raw).suffix.lower()
    if suffix not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"Unsupported type: {suffix}")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}{suffix}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(save_path, "wb") as f: f.write(content)
    text = ""
    if suffix == ".pdf": text = _parse_pdf(save_path)
    elif suffix == ".docx": text = _parse_docx(save_path)
    elif suffix == ".xlsx": text = _parse_xlsx(save_path)
    return ApiResponse(data={"filename": safe_name, "original_name": safe_name_raw, "size": len(content), "extracted_text": text, "url": f"/uploads/{safe_name}"})


# ═══════════════════════════════════════════════════════════════
# 视频/音频 SSE 流式导入（带实时进度）
# ═══════════════════════════════════════════════════════════════

def _sse_fmt(event: str, data: dict) -> str:
    """格式化 SSE 事件"""
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def _fmt_dur(seconds: float) -> str:
    """秒 → m分s秒"""
    m, s = divmod(int(seconds), 60)
    if m > 0:
        return f"{m}分{s}秒"
    return f"{s}秒"


@router.post("/upload/video-import-stream")
async def video_import_stream(
    request: Request,
    file: UploadFile = File(...),
    category_id: int = Form(0),
    knowledge_base: str = Form("public"),
    user: User = Depends(require_admin),
):
    """视频智能导入 SSE 流 —— 实时推送每一步进度和预估剩余时间"""

    # ── 在生成器外部完成文件保存（避免生成器内 SpooledTemporaryFile 被关闭） ──
    safe_name_raw = _safe_filename(file.filename)
    suffix = Path(safe_name_raw).suffix.lower()
    if suffix not in (".mp4", ".webm", ".mp3", ".wav"):
        raise HTTPException(status_code=400, detail=f"不支持的文件类型: {suffix}")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}{suffix}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    file_size_mb = len(content) / 1024 / 1024
    with open(save_path, "wb") as f:
        f.write(content)
    base_title = Path(safe_name_raw).stem
    # 在 session 有效期内读取 user.dept（避免生成器内 DetachedInstanceError）
    source_dept = user.dept.name if user.dept else ""
    # 预加载 ASR 密钥（避免 SSE 生成器内 async_session 嵌套问题）
    import asyncio as _asyncio2
    try:
        from asr import _load_keys_from_db as _asr_load, reset_asr_cache as _asr_reset
        _asr_reset()  # 清除可能缓存的空值
        _asr_keys = await _asyncio2.ensure_future(_asr_load())
    except Exception:
        _asr_keys = ("", "")
    logger.info(f"[SSE] file saved: {safe_name} ({file_size_mb:.1f}MB), ASR keys={'ready' if _asr_keys[0] else 'none'}, starting SSE stream")

    async def generate():
        t0 = _time.time()

        # ── Step 1: 文件已保存 ──
        yield _sse_fmt("progress", {
            "step": "save_done", "detail": f"文件已保存 ({file_size_mb:.1f}MB)",
            "pct": 5, "elapsed_sec": _time.time() - t0,
        })

        # ── Step 2: 获取视频时长 + 预估总耗时 ──
        from video_processor import (
            _extract_audio_async, _transcribe_with_timestamps, _create_segments,
            _create_single_entry, _get_video_duration, _format_ts,
        )
        vid_dur = await _get_video_duration(save_path)
        # 预估公式：ffmpeg(dur*0.25) + ASR(dur*0.7) + AI(15s) + 开销(8s)
        est_total = max(vid_dur * 0.95 + 23, 20)
        yield _sse_fmt("progress", {
            "step": "estimate", "detail": f"视频时长 {_fmt_dur(vid_dur)}，预估 {_fmt_dur(est_total)} 完成",
            "pct": 8, "estimated_total_sec": round(est_total, 1),
            "video_duration_sec": round(vid_dur, 1),
        })

        # ── Step 3: 匹配分类 ──
        yield _sse_fmt("progress", {"step": "category", "detail": "匹配知识分类...", "pct": 10,
                                    "elapsed_sec": _time.time() - t0})
        matched_cat = None
        async with async_session() as db:
            if category_id > 0:
                r = await db.execute(select(KnowledgeCategory).where(KnowledgeCategory.id == category_id))
                matched_cat = r.scalar_one_or_none()
            if not matched_cat:
                r = await db.execute(
                    select(KnowledgeCategory).where(KnowledgeCategory.knowledge_base == knowledge_base)
                    .order_by(KnowledgeCategory.sort_order).limit(1)
                )
                matched_cat = r.scalar_one_or_none()
            if not matched_cat:
                r = await db.execute(select(KnowledgeCategory).order_by(KnowledgeCategory.sort_order).limit(1))
                matched_cat = r.scalar_one_or_none()
        if not matched_cat:
            yield _sse_fmt("error", {"detail": "无可用知识分类，请先创建分类"})
            return
        cat_info = {"id": matched_cat.id, "name": matched_cat.name,
                    "knowledge_base": matched_cat.knowledge_base.value}
        yield _sse_fmt("progress", {
            "step": "category_done", "detail": f"分类: {matched_cat.icon or '📄'} {matched_cat.name}",
            "pct": 12, "category": cat_info,
        })

        # ── Step 4: 提取音频（后台异步任务 + 循环 yield 进度）──
        yield _sse_fmt("progress", {
            "step": "audio_extract", "detail": "正在提取音频（ffmpeg）...",
            "pct": 15, "elapsed_sec": round(_time.time() - t0, 1),
        })
        audio_path = os.path.join(tempfile.gettempdir(), f"vocal_{uuid.uuid4().hex[:8]}.mp3")
        import asyncio as _asyncio
        _progress_q: _asyncio.Queue = _asyncio.Queue()
        def _audio_cb(pct: float, detail: str):
            _progress_q.put_nowait((pct, detail))
        _audio_task = _asyncio.ensure_future(
            _extract_audio_async(save_path, audio_path, on_progress=_audio_cb)
        )
        try:
            last_fwd_pct = 15
            while not _audio_task.done():
                # 排空进度队列
                while not _progress_q.empty():
                    pct, detail = _progress_q.get_nowait()
                    mapped = 15 + int(pct * 0.15)
                    if mapped > last_fwd_pct + 1:
                        last_fwd_pct = mapped
                        yield _sse_fmt("progress", {
                            "step": "audio_extract", "detail": detail,
                            "pct": min(mapped, 29),
                            "elapsed_sec": round(_time.time() - t0, 1),
                        })
                await _asyncio.sleep(0.5)  # 0.5s 轮询，不烧 CPU
            # 清空剩余事件
            while not _progress_q.empty():
                pct, detail = _progress_q.get_nowait()
                mapped = 15 + int(pct * 0.15)
                if mapped > last_fwd_pct:
                    last_fwd_pct = mapped
                    yield _sse_fmt("progress", {
                        "step": "audio_extract", "detail": detail,
                        "pct": min(mapped, 29),
                        "elapsed_sec": round(_time.time() - t0, 1),
                    })
            await _audio_task  # 拿到异常（如果有）
            yield _sse_fmt("progress", {
                "step": "audio_done", "detail": "音频提取完成 ✓",
                "pct": 30, "elapsed_sec": round(_time.time() - t0, 1),
            })
        except Exception as e:
            logger.warning(f"音频提取失败: {e}")
            yield _sse_fmt("progress", {
                "step": "audio_fail",
                "detail": f"音频提取失败: {str(e)[:60]}，创建单条占位条目",
                "pct": 90,
            })
            try:
                eids = await _create_single_entry(
                    base_title, "", matched_cat.id, matched_cat.knowledge_base.value,
                    user.real_name, source_dept, f"批量导入,{matched_cat.name}",
                    save_path, content_type="video",
                )
                yield _sse_fmt("done", {"entry_ids": eids, "drafts": [],
                              "total_sec": round(_time.time() - t0, 1),
                              "category": cat_info})
            except Exception as e2:
                yield _sse_fmt("error", {"detail": f"创建条目失败: {e2}"})
            return

        # ── Step 5: ASR 转写 ──
        yield _sse_fmt("progress", {
            "step": "transcribe", "detail": "正在语音转文字（ASR）... 这是最慢的步骤",
            "pct": 35, "elapsed_sec": round(_time.time() - t0, 1),
        })
        segments = await _transcribe_with_timestamps(audio_path)
        el = round(_time.time() - t0, 1)
        if not segments:
            # ASR 全不可用 → 静态分段 pending
            yield _sse_fmt("progress", {
                "step": "transcribe_fallback",
                "detail": "无可用 ASR，使用静态分段（pending 待人工填写）",
                "pct": 60, "elapsed_sec": el,
            })
            from video_processor import _get_audio_duration, _static_segments
            dur = await _get_audio_duration(audio_path)
            segments = _static_segments(dur)
            seg_status = EntryStatusEnum.pending
        else:
            yield _sse_fmt("progress", {
                "step": "transcribe_done",
                "detail": f"转写完成：{len(segments)} 段，{sum(len(s['text']) for s in segments)} 字",
                "pct": 65, "elapsed_sec": el,
            })
            seg_status = EntryStatusEnum.approved

        # ── Step 6: 分段入库 ──
        yield _sse_fmt("progress", {
            "step": "segments", "detail": f"正在写入 {len(segments)} 个知识片段...",
            "pct": 72,
        })
        try:
            entry_ids = await _create_segments(
                base_title, segments, matched_cat.id, matched_cat.knowledge_base.value,
                user.real_name, source_dept, f"批量导入,{matched_cat.name}",
                save_path, status=seg_status,
            )
            yield _sse_fmt("progress", {
                "step": "segments_done", "detail": f"已生成 {len(entry_ids)} 个知识片段",
                "pct": 82, "entry_ids": entry_ids,
                "elapsed_sec": round(_time.time() - t0, 1),
            })
        except Exception as e:
            yield _sse_fmt("error", {"detail": f"写入知识条目失败: {e}"})
            return

        # ── Step 7: AI 出题（仅当有真实 ASR 转写结果时才执行） ──
        use_real_asr = seg_status != EntryStatusEnum.pending  # 静态分段跳过 AI 出题
        if not use_real_asr:
            yield _sse_fmt("progress", {
                "step": "ai_skip",
                "detail": "无 ASR 转写内容，跳过 AI 出题（可手动编辑片段内容后再点击AI拆分按钮）",
                "pct": 95, "elapsed_sec": round(_time.time() - t0, 1),
            })
            yield _sse_fmt("done", {
                "entry_ids": entry_ids, "drafts": [],
                "total_sec": round(_time.time() - t0, 1),
                "category": cat_info,
                "message": f"视频处理完成：{len(entry_ids)} 个片段（待填写内容后手动AI出题）",
            })
            return

        yield _sse_fmt("progress", {
            "step": "ai_questions", "detail": "正在调用 AI 拆解试题...",
            "pct": 87, "elapsed_sec": round(_time.time() - t0, 1),
        })
        drafts = []
        combined_text = ""
        async with async_session() as db:
            for eid in entry_ids:
                r = await db.execute(select(KnowledgeEntry).where(KnowledgeEntry.id == eid))
                ke = r.scalar_one_or_none()
                if ke and ke.content:
                    combined_text += ke.content + "\n\n"

        if combined_text.strip():
            tpos = matched_cat.knowledge_base.value if matched_cat.knowledge_base.value in ("sales", "tech", "service") else ""
            try:
                drafts = await _ai_gen(combined_text[:8000], tpos)
                if drafts:
                    drafts = _normalize_drafts(drafts)
                    yield _sse_fmt("progress", {
                        "step": "ai_done", "detail": f"AI 已生成 {len(drafts)} 道试题草稿",
                        "pct": 97, "drafts_count": len(drafts),
                        "elapsed_sec": round(_time.time() - t0, 1),
                    })
                else:
                    yield _sse_fmt("progress", {
                        "step": "ai_fail", "detail": "AI 出题失败(LLM 未配置或生成失败)",
                        "pct": 97,
                    })
            except Exception as e:
                yield _sse_fmt("progress", {
                    "step": "ai_fail", "detail": f"AI 出题异常: {str(e)[:80]}",
                    "pct": 97,
                })
        else:
            # 纯静态分段没有实际文本
            yield _sse_fmt("progress", {
                "step": "ai_skip", "detail": "无转写文本，跳过 AI 出题（请先人工填写转写内容）",
                "pct": 97,
            })

        # ── Done ──
        total_s = round(_time.time() - t0, 1)
        yield _sse_fmt("done", {
            "entry_ids": entry_ids,
            "drafts": drafts,
            "total_sec": total_s,
            "category": cat_info,
            "message": f"视频处理完成：{len(entry_ids)} 个片段" +
                       (f"，{len(drafts)} 道题目草稿" if drafts else ""),
        })

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "X-Accel-Buffering": "no",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        },
    )


@router.post("/upload/smart-import", response_model=ApiResponse)
async def smart_import(
    file: UploadFile = File(...),
    category_id: int = Form(0),
    knowledge_base: str = Form("public"),
    question_count: int = Form(0),
    user: User = Depends(require_admin),
):
    steps = []
    result = {"knowledge_id": None, "question_ids": [], "category_matched": None}

    # Step 1: save file
    safe_name_raw = _safe_filename(file.filename)
    suffix = Path(safe_name_raw).suffix.lower()
    if suffix not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型：{suffix}")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}{suffix}"
    save_path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(save_path, "wb") as f: f.write(content)
    steps.append({"step": "file_saved", "status": "ok", "detail": f"文件已保存: {safe_name}"})

    # Step 2: extract text
    text = ""
    engine = ""
    if suffix == ".pdf":
        text = _parse_pdf(save_path); engine = "PDF解析"
    elif suffix == ".docx":
        text = _parse_docx(save_path); engine = "DOCX解析"
    elif suffix == ".xlsx":
        text = _parse_xlsx(save_path); engine = "XLSX解析"
    tlen = len(text)
    if tlen < 20:
        steps.append({"step": "text_extract", "status": "fail", "detail": f"{engine}失败：文本过短({tlen}字符)"})
        return ApiResponse(code=400, data={**result, "steps": steps}, msg="text extraction failed")
    steps.append({"step": "text_extract", "status": "ok", "detail": f"{engine}完成，提取 {tlen} 字符"})

    # ── 视频/音频分支：不走文本流程，委托 video_processor ──
    if suffix in (".mp4", ".webm", ".mp3", ".wav"):
        from video_processor import process_video
        # 分类：优先用传入的 category_id，否则按 knowledge_base 取第一个匹配分类
        vid_cat = None
        async with async_session() as db:
            if category_id > 0:
                cat_r = await db.execute(select(KnowledgeCategory).where(KnowledgeCategory.id == category_id))
                vid_cat = cat_r.scalar_one_or_none()
            if not vid_cat:
                cat_r = await db.execute(
                    select(KnowledgeCategory).where(KnowledgeCategory.knowledge_base == knowledge_base)
                    .order_by(KnowledgeCategory.sort_order).limit(1)
                )
                vid_cat = cat_r.scalar_one_or_none()
            if not vid_cat:
                cat_r = await db.execute(select(KnowledgeCategory).order_by(KnowledgeCategory.sort_order).limit(1))
                vid_cat = cat_r.scalar_one_or_none()

        if not vid_cat:
            steps.append({"step": "category_match", "status": "fail", "detail": "无可用分类"})
            return ApiResponse(code=400, data={**result, "steps": steps}, msg="请先创建知识分类")

        steps.append({"step": "category_match", "status": "ok",
                      "detail": f"分类: {vid_cat.icon or '📄'} {vid_cat.name}"})
        result["category_matched"] = {"id": vid_cat.id, "name": vid_cat.name, "knowledge_base": vid_cat.knowledge_base.value}

        # 视频转写 + 分段入库
        try:
            entry_ids = await process_video(
                video_path=save_path,
                category_id=vid_cat.id,
                knowledge_base=vid_cat.knowledge_base.value,
                title=Path(safe_name_raw).stem,
                source_person=user.real_name,
                source_dept=user.dept.name if user.dept else "",
                tags=f"批量导入,{vid_cat.name}",
            )
            if entry_ids:
                result["knowledge_id"] = entry_ids[0]
                steps.append({"step": "knowledge_saved", "status": "ok",
                              "detail": f"视频转写完成，生成 {len(entry_ids)} 个知识片段"})
            else:
                steps.append({"step": "knowledge_saved", "status": "fail",
                              "detail": "视频处理未产生知识条目"})
        except Exception as e:
            steps.append({"step": "knowledge_saved", "status": "fail",
                          "detail": f"视频处理失败: {e}"})
            return ApiResponse(code=500, data={**result, "steps": steps}, msg=f"视频处理失败: {e}")

        # 合并片段文本 → AI 出题
        drafts = []
        if entry_ids and question_count > 0:
            combined_text = ""
            async with async_session() as db:
                for eid in entry_ids:
                    ke_r = await db.execute(select(KnowledgeEntry).where(KnowledgeEntry.id == eid))
                    ke = ke_r.scalar_one_or_none()
                    if ke and ke.content:
                        combined_text += ke.content + "\n\n"
            if combined_text.strip():
                tpos = vid_cat.knowledge_base.value if vid_cat.knowledge_base.value in ("sales", "tech", "service") else ""
                try:
                    drafts = await _ai_gen(combined_text[:8000], tpos)
                    if drafts:
                        drafts = _normalize_drafts(drafts)
                        steps.append({"step": "ai_questions", "status": "ok",
                                      "detail": f"AI拆解完成，生成 {len(drafts)} 道题目草稿"})
                    else:
                        steps.append({"step": "ai_questions", "status": "fail",
                                      "detail": "AI出题失败(LLM未配置或生成失败)"})
                except Exception as e:
                    steps.append({"step": "ai_questions", "status": "fail", "detail": f"AI出题异常: {e}"})

        return ApiResponse(
            data={**result, "steps": steps, "drafts": drafts},
            msg=f"视频智能导入完成: {len(entry_ids or [])} 个转写片段" +
                (f"+{len(drafts)}道题目草稿" if drafts else "")
        )

    # ── 以下为原有 PDF/DOCX/XLSX 流程 ──
    # Step 2b: 检测是否已是结构化试题 → 直接解析，跳过AI拆题
    parsed_drafts = _detect_and_parse(text)
    if parsed_drafts:
        steps.append({"step": "structured_detect", "status": "ok",
                      "detail": f"检测到已编排的结构化试题 {len(parsed_drafts)} 道，直接解析，无需AI拆分"})
        return ApiResponse(
            data={**result, "steps": steps, "extracted_text": text, "text_length": tlen,
                  "drafts": parsed_drafts, "direct_parse": True},
            msg=f"检测到结构化试题 {len(parsed_drafts)} 道，已直接解析，请复核后入库")

    # Step 3: match category
    matched_cat = None
    async with async_session() as db:
        cats = (await db.execute(
            select(KnowledgeCategory).order_by(KnowledgeCategory.knowledge_base, KnowledgeCategory.sort_order)
        )).scalars().all()
        if category_id > 0:
            for c in cats:
                if c.id == category_id:
                    matched_cat = c; break
            if matched_cat:
                steps.append({"step": "category_match", "status": "ok",
                              "detail": f"指定分类: {matched_cat.icon} {matched_cat.name}"})
            else:
                steps.append({"step": "category_match", "status": "skip",
                              "detail": "指定分类无效，跳过入库和拆题"})
                return ApiResponse(code=200, data={**result, "steps": steps, "extracted_text": text},
                                   msg="分类无效，未执行后续操作")
        else:
            # auto match by keyword scoring
            best, best_score = None, 0
            for c in cats:
                sc = 0
                if c.name in text: sc += 10
                for ch in c.name:
                    if ch in text: sc += 1
                if sc > best_score: best, best_score = c, sc
            if best and best_score >= 3:
                matched_cat = best
                steps.append({"step": "category_match", "status": "ok",
                              "detail": f"自动匹配: {best.icon} {best.name} (得分{best_score})"})
            else:
                steps.append({"step": "category_match", "status": "skip",
                              "detail": f"自动匹配失败(最高得分{best_score})，未执行入库和拆题"})
                return ApiResponse(code=200, data={**result, "steps": steps, "extracted_text": text},
                                   msg="分类匹配度不足，未执行入库和拆题")

    result["category_matched"] = {"id": matched_cat.id, "name": matched_cat.name, "knowledge_base": matched_cat.knowledge_base.value}

    # Step 4: knowledge entry
    try:
        async with async_session() as db:
            ke = KnowledgeEntry(
                title=Path(safe_name_raw).stem, content=text[:10000],
                content_type=ContentTypeEnum.text, category_id=matched_cat.id,
                knowledge_base=matched_cat.knowledge_base.value,
                source_type=SourceTypeEnum.manual, source_file_path=save_path,
                source_person=user.real_name, status=EntryStatusEnum.approved,
                tags=f"批量导入,{matched_cat.name}",
            )
            db.add(ke); await db.commit(); await db.refresh(ke)
            result["knowledge_id"] = ke.id
            steps.append({"step": "knowledge_saved", "status": "ok",
                          "detail": f"知识已入库: ID={ke.id} ({ke.title})"})
    except Exception as e:
        steps.append({"step": "knowledge_saved", "status": "fail", "detail": f"知识入库失败: {e}"})

    # Step 5: AI question generation (AI自行判断出题数量)
    tpos = ""
    if matched_cat.knowledge_base.value in ("sales", "tech", "service"):
        tpos = matched_cat.knowledge_base.value
    drafts = []
    if question_count > 0:
        try:
            drafts = await _ai_gen(text, tpos)
            if drafts:
                drafts = _normalize_drafts(drafts)
                steps.append({"step": "ai_questions", "status": "ok",
                              "detail": f"AI拆解完成，生成 {len(drafts)} 道题目草稿（待人工复核）"})
            else:
                steps.append({"step": "ai_questions", "status": "fail", "detail": "AI出题失败(LLM未配置或生成失败)"})
        except Exception as e:
            steps.append({"step": "ai_questions", "status": "fail", "detail": f"AI出题异常: {e}"})

    return ApiResponse(data={**result, "steps": steps, "extracted_text": text, "text_length": tlen, "drafts": drafts},
                       msg=f"智能导入完成: 知识已入库+{len(drafts)}道题目草稿待复核")


async def _ai_gen(text: str, target: str) -> list:
    async with async_session() as db:
        llm = (await db.execute(
            select(LLMProvider).where(LLMProvider.is_active == True, LLMProvider.is_default == True)
        )).scalar_one_or_none()
    if not llm or not llm.api_key: return []
    key = LLM_ENCRYPTION_KEY.encode()
    digest = hashlib.sha256(key).digest()
    b64 = base64.urlsafe_b64encode(digest)
    api_key = Fernet(b64).decrypt(llm.api_key.encode()).decode()
    pos_label = {"sales":"销售","tech":"技术","service":"客服"}.get(target,"通用")
    prompt = (
        f"你是合群汽车集团出题专家。阅读以下文档，根据文档的知识点密度和内容丰富度，自行判断生成适当数量的考题（内容少的少出、知识点密集的多出，建议3-20道）。\n"
        f"目标岗位：{pos_label}。不要为了凑数生成重复或浅显的题目。\n"
        f"文档：\n{text[:8000]}\n\n"
        f"输出JSON数组，每题格式：{{\"question_type\":\"single_choice\",\"question_content\":\"题干\","
        f"\"options\":{{\"A\":\"A\",\"B\":\"B\",\"C\":\"C\",\"D\":\"D\"}},\"answer\":\"A\","
        f"\"explanation\":\"解析\",\"difficulty_level\":2}}。只输出JSON数组。"
    )
    try:
        async with httpx.AsyncClient(timeout=60) as c:
            resp = await c.post(f"{llm.base_url}/chat/completions",
                headers={"Authorization": f"Bearer {api_key}"},
                json={"model": llm.model_name, "messages": [
                    {"role":"system","content":"你是出题助手，只输出JSON数组。"},
                    {"role":"user","content":prompt},
                ], "temperature":0.8, "max_tokens":8192})
            resp.raise_for_status()
            t = resp.json()["choices"][0]["message"]["content"].strip()
            if t.startswith("```"): t = t.split("\n",1)[1].rsplit("\n```",1)[0]
            if t.startswith("```json"): t = t.split("\n",1)[1].rsplit("\n```",1)[0]
            return json.loads(t)
    except Exception as e:
        logger.warning(f"AI出题失败: {e}")
        return []


# === Structured question detection ===

def _detect_and_parse(text: str) -> list | None:
    """检测文本是否已是结构化试题。是则解析返回题目列表，否则返回 None。"""
    t = text.strip()
    if not t:
        return None

    # 方法1：JSON 数组格式
    if t.startswith("["):
        try:
            parsed = json.loads(t)
            if isinstance(parsed, list) and len(parsed) > 0:
                if all(isinstance(q, dict) and "question_content" in q and "answer" in q for q in parsed):
                    logger.info(f"检测到 JSON 格式试题 {len(parsed)} 道")
                    return _normalize_drafts(parsed)
        except (json.JSONDecodeError, Exception):
            pass

    # 方法2：中文试题格式（题号 + ABCD选项 + 答案）
    has_numbers = bool(re.search(r'(?:^|\n)\s*\d+[\.、）\)]\s*\S', t))
    has_options = bool(re.search(r'[A-D][\.、．）\)]', t))
    has_answer = bool(re.search(r'(?:正确答案|答案|answer)[：:]\s*[A-D]', t, re.IGNORECASE))

    if has_numbers and has_options and has_answer:
        questions = _parse_text_questions(t)
        if questions:
            logger.info(f"检测到文本格式试题 {len(questions)} 道")
            return _normalize_drafts(questions)

    return None


def _normalize_drafts(qs: list) -> list:
    """补全草稿默认字段，确保前端可用。"""
    for q in qs:
        if "question_type" not in q or q["question_type"] not in ("single_choice","multi_choice","true_false","fill_blank"):
            q["question_type"] = "single_choice"
        if "difficulty_level" not in q:
            q["difficulty_level"] = 2
        if "options" not in q or not isinstance(q.get("options"), dict):
            # 尝试从 opt_A/opt_B 或 A/B 键容错
            opts = {}
            for k in ("A","B","C","D"):
                val = q.pop(f"opt_{k}", None) or q.pop(k, None)
                if val:
                    opts[k] = str(val)
            q["options"] = opts if opts else {"A":"","B":"","C":"","D":""}
        if "explanation" not in q:
            q["explanation"] = ""
        q["question_content"] = str(q.get("question_content", ""))[:2000]
        q["answer"] = str(q.get("answer", "A"))[:500]
    return qs


def _parse_text_questions(text: str) -> list:
    """将中文试题文本拆成题目列表。"""
    questions = []
    # 按题号切块
    blocks = re.split(r'\n(?=\d+[\.、）\)]\s+\S)', text)
    if len(blocks) <= 1:
        # 尝试 Q1/Q2 格式
        blocks = re.split(r'\n(?=Q\d+\s*[：:])', text, flags=re.IGNORECASE)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        # 题干
        qm = re.search(r'(?:\d+[\.、）\)]|Q\d+\s*[：:])\s*(.+?)(?=\n\s*[A-D][\.、．）\)]|\n\s*(?:正确答案|答案|answer)|$)', block, re.IGNORECASE | re.DOTALL)
        question_content = qm.group(1).strip() if qm else ""

        # 选项
        options = {}
        for om in re.finditer(r'([A-D])[\.、．）\)]\s*(.+?)(?=\n\s*[A-D][\.、．）\)]|\n\s*(?:正确答案|答案|answer|解析|解释)|\Z)', block, re.IGNORECASE | re.DOTALL):
            options[om.group(1)] = om.group(2).strip()

        # 答案
        am = re.search(r'(?:正确答案|答案|answer)[：:]\s*([A-D])', block, re.IGNORECASE)
        answer = am.group(1) if am else ""

        # 解析
        em = re.search(r'(?:解析|解释|explanation)[：:]\s*(.+?)(?=\n\d+[\.、）\)]|\Z)', block, re.IGNORECASE | re.DOTALL)
        explanation = em.group(1).strip() if em else ""

        if question_content and answer:
            questions.append({
                "question_type": "single_choice",
                "question_content": question_content,
                "options": options if options else {"A":"","B":"","C":"","D":""},
                "answer": answer,
                "explanation": explanation,
                "difficulty_level": 2,
            })

    return questions


# === Parsers ===

def _parse_pdf(path: str) -> str:
    text = ""; tried = []
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path); tried.append("PyPDF2")
        if reader.is_encrypted: return "[PDF已加密]"
        for page in reader.pages:
            t = page.extract_text()
            if t and t.strip(): text += t + "\n"
        if text.strip(): return text.strip()
    except Exception: pass
    try:
        import pdfplumber; tried.append("pdfplumber")
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t and t.strip(): text += t + "\n"
        if text.strip(): return text.strip()
    except Exception: pass
    try:
        tried.append("RapidOCR")
        ocr = _rapidocr_pdf(path)
        if ocr.strip(): return ocr.strip()
    except Exception: pass
    return f"[识别失败] 已尝试{'、'.join(tried)}均未获得文字。文件已保存。"


def _rapidocr_pdf(path: str) -> str:
    import numpy as np
    from PyPDF2 import PdfReader
    from PIL import Image
    ocr = _get_ocr(); reader = PdfReader(path); all_texts = []
    for pi, page in enumerate(reader.pages):
        img_arr = None
        try:
            resources = page["/Resources"]
            if "/XObject" in resources:
                for key in resources["/XObject"]:
                    try:
                        obj = resources["/XObject"][key]
                        if obj["/Subtype"] == "/Image":
                            raw = obj.get_data()
                            img = Image.open(io.BytesIO(raw))
                            if img.mode != "RGB": img = img.convert("RGB")
                            img_arr = np.array(img); break
                    except Exception: continue
        except Exception: continue
        if img_arr is None: continue
        try:
            result, _ = ocr(img_arr)
            if result:
                lines = []
                for item in result:
                    txt = item[1]
                    try: conf = float(item[2])
                    except: conf = 0.0
                    if txt and conf > 0.3: lines.append(txt)
                pt = " ".join(lines).strip()
                if pt: all_texts.append(f"--- 第{pi+1}页 ---\n{pt}")
        except Exception: pass
    return "\n\n".join(all_texts)


def _parse_docx(path: str) -> str:
    try:
        from docx import Document
        return "\n".join(p.text for p in Document(path).paragraphs if p.text.strip())
    except Exception as e: return f"[DOCX错误:{e}]"


def _parse_xlsx(path: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        rows = []
        for sn in wb.sheetnames:
            ws = wb[sn]; headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: headers = [str(h) if h else f"C{j}" for j, h in enumerate(row)]
                elif any(c is not None for c in row):
                    parts = [f"{headers[j]}:{cell}" for j, cell in enumerate(row) if cell is not None and j < len(headers)]
                    rows.append(";".join(parts))
        return "\n".join(rows[:500])
    except Exception as e: return f"[XLSX错误:{e}]"
